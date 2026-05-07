from __future__ import annotations

import csv
import json
import re
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


ROOT = Path(__file__).resolve().parents[1]
OUTPUTS_DIR = ROOT / "outputs"
EXPORT_DIR = OUTPUTS_DIR / "existing_workbook_export"
WORKBOOK_OUT = OUTPUTS_DIR / "zepto_expense_split_ready.xlsx"
BASELINE_MANIFEST_PATH = EXPORT_DIR / "summary.json"

BASELINE_SHEET_NAMES = {
    "1_Orders_Summary": "Orders Summary",
    "2_Line_Items": "Line Items",
    "3_Charges_Taxes": "Charges Taxes",
    "4_Monthly_Summary": "Monthly Summary",
    "5_Category_Summary": "Category Summary",
    "6_Vendor_Summary": "Vendor Summary",
    "7_Data_Quality": "Data Quality",
}

CATEGORY_VOCABULARY = [
    "groceries",
    "household",
    "personal",
    "medicines",
    "snacks",
    "misc",
]

CATEGORY_KEYWORDS = {
    "groceries": [
        "milk",
        "paneer",
        "tomato",
        "potato",
        "onion",
        "vegetable",
        "fruit",
        "curd",
        "bread",
        "eggs",
        "egg",
        "mushroom",
        "rice",
        "atta",
        "dal",
        "hybrid",
    ],
    "household": [
        "cleaner",
        "detergent",
        "mop",
        "tissue",
        "foil",
        "garbage bag",
        "storage",
        "container",
        "basket",
        "dishwash",
    ],
    "personal": [
        "marlboro",
        "cigarette",
        "grooming",
        "cosmetic",
        "hydration drink",
        "zero sugar",
        "deodorant",
        "shampoo",
        "soap",
        "razor",
    ],
    "medicines": [
        "vicks",
        "medicine",
        "tablet",
        "syrup",
        "ointment",
        "capsule",
        "capsules",
        "pharmacy",
        "melatonin",
        "sleep gummies",
        "gummies",
        "healthcare",
    ],
    "snacks": [
        "potato chips",
        "chips",
        "puffs",
        "cookies",
        "cookie",
        "chocolate",
        "candy",
        "soft drink",
        "biscuit",
        "snack",
        "kurkure",
        "lays",
    ],
}

PROMO_PATTERNS = [
    "flyer",
    "tv flyer",
    "hope starts here",
]


def clean_cell(value: Any) -> str:
    return (
        str(value if value is not None else "")
        .replace("\r", " ")
        .replace("\n", " | ")
        .replace("  ", " ")
        .strip()
    )


def read_json(file_path: Path, fallback: Any) -> Any:
    try:
        return json.loads(file_path.read_text(encoding="utf-8"))
    except FileNotFoundError:
        return fallback


def read_csv(file_path: Path) -> list[list[str]]:
    with file_path.open(newline="", encoding="utf-8") as handle:
        return [list(row) for row in csv.reader(handle)]


def source_order_id_from_file_name(file_name: Any) -> str:
    match = re.match(
        r"^([0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12})-",
        str(file_name or ""),
        re.IGNORECASE,
    )
    return match.group(1) if match else ""


def line_item_key_for(order_id: str, record: dict[str, Any], index: int) -> str:
    return f"{order_id or ''}::{record.get('invoice_number') or ''}::{index}"


def has_own(record: dict[str, Any] | None, key: str) -> bool:
    return isinstance(record, dict) and key in record


def value_from_annotation(
    line_annotation: dict[str, Any],
    order_annotation: dict[str, Any],
    key: str,
    fallback: Any = "",
) -> Any:
    if has_own(line_annotation, key):
        return line_annotation[key]
    if has_own(order_annotation, key):
        return order_annotation[key]
    return fallback


def owns_any_annotation_field(record: dict[str, Any], fields: list[str]) -> bool:
    return any(has_own(record, field) for field in fields)


def split_tag_source(line_annotation: dict[str, Any], order_annotation: dict[str, Any]) -> str:
    tag_fields = ["expense_category", "split_type", "split_with", "ready_for_splitwise"]
    if owns_any_annotation_field(line_annotation, tag_fields):
        return "line_item"
    if owns_any_annotation_field(order_annotation, tag_fields):
        return "order"
    return "suggested"


def review_status_from_annotation(
    line_annotation: dict[str, Any],
    order_annotation: dict[str, Any],
) -> str:
    return str(value_from_annotation(line_annotation, order_annotation, "review_status", "unreviewed"))


def normalize_text(value: Any) -> str:
    return str(value or "").strip().lower()


def round_currency(value: Any) -> float:
    try:
        return round(float(value or 0), 2)
    except (TypeError, ValueError):
        return 0


def parse_amount(value: Any) -> float:
    if isinstance(value, (int, float)):
        return round_currency(value)
    text = str(value or "").strip()
    if not text:
        return 0
    normalized = text.replace(",", "")
    multiplication = re.search(
        r"(\d+(?:\.\d+)?)\s*(?:x|\u00d7)\s*(?:rs\.?\s*)?(\d+(?:\.\d+)?)",
        normalized,
        re.IGNORECASE,
    )
    if multiplication:
        return round_currency(float(multiplication.group(1)) * float(multiplication.group(2)))
    matches = re.findall(r"-?\d+(?:\.\d+)?", normalized)
    if not matches:
        return 0
    decimal_matches = [match for match in matches if "." in match]
    selected = decimal_matches[-1] if decimal_matches else matches[-1]
    return round_currency(selected)


def find_phrase_ranges(text: str, phrase: str) -> list[tuple[int, int]]:
    parts = [re.escape(part) for part in phrase.split() if part]
    if not parts:
        return []
    phrase_pattern = r"[\s./_-]+".join(parts)
    pattern = re.compile(rf"(^|[^a-z0-9])({phrase_pattern})(?=[^a-z0-9]|$)")
    ranges: list[tuple[int, int]] = []
    for match in pattern.finditer(text):
        prefix = match.group(1) or ""
        start = match.start() + len(prefix)
        ranges.append((start, start + len(match.group(2) or "")))
    return ranges


def is_promo_like(text: str) -> bool:
    return any(find_phrase_ranges(text, pattern) for pattern in PROMO_PATTERNS)


def find_keyword_matches(text: str, keywords: list[str]) -> list[str]:
    occupied: list[tuple[int, int]] = []
    matches: list[str] = []
    for keyword in sorted(keywords, key=lambda item: (-len(item), item)):
        for phrase_range in find_phrase_ranges(text, keyword):
            overlaps = any(phrase_range[0] < end and phrase_range[1] > start for start, end in occupied)
            if overlaps:
                continue
            occupied.append(phrase_range)
            matches.append(keyword)
            break
    return matches


def collect_row_signals(row: dict[str, Any]) -> dict[str, Any]:
    text = normalize_text(row.get("item_description"))
    amount = parse_amount(row.get("line_total_amount"))
    promo_like = is_promo_like(text)
    multiplier = 0.02 if promo_like else 1
    matches_by_category = {}
    for category in CATEGORY_VOCABULARY:
        if category == "misc":
            continue
        matches = find_keyword_matches(text, CATEGORY_KEYWORDS.get(category, []))
        if matches:
            matches_by_category[category] = matches
    return {
        "amount": amount,
        "matchesByCategory": matches_by_category,
        "promoLike": promo_like,
        "text": text,
        "weightedAmount": amount * multiplier,
    }


def determine_confidence(winner_score: float, runner_up_score: float) -> str:
    if winner_score >= 75 and winner_score >= (runner_up_score * 2 or 1):
        return "high"
    if winner_score >= 20 and winner_score > runner_up_score:
        return "medium"
    return "low"


def suggest_order_category(order: dict[str, Any]) -> dict[str, Any]:
    scores = {category: 0.0 for category in CATEGORY_VOCABULARY}
    tie_breakers = {category: 0 for category in CATEGORY_VOCABULARY}
    reasons_by_category: dict[str, list[dict[str, Any]]] = {category: [] for category in CATEGORY_VOCABULARY}

    for row in order.get("invoice_rows") or []:
        signals = collect_row_signals(row)
        matched_categories = list(signals["matchesByCategory"].items())
        row_score = signals["weightedAmount"] / len(matched_categories) if matched_categories else 0
        for category, matches in matched_categories:
            scores[category] += row_score
            tie_breakers[category] += max((len(keyword) for keyword in matches), default=0)
            for keyword in matches:
                reasons_by_category[category].append({"keyword": keyword, "score": row_score})

    ranked = sorted(
        (
            {"category": category, "score": scores[category], "tieBreaker": tie_breakers[category]}
            for category in CATEGORY_VOCABULARY
            if category != "misc"
        ),
        key=lambda row: (-row["score"], -row["tieBreaker"], row["category"]),
    )
    winner = ranked[0] if ranked else {"category": "misc", "score": 0}
    runner_up = ranked[1] if len(ranked) > 1 else {"score": 0}

    if not winner["score"]:
        return {"category": "misc", "confidence": "low", "reasons": []}

    reasons: list[str] = []
    for entry in sorted(
        reasons_by_category[winner["category"]],
        key=lambda row: (-row["score"], row["keyword"]),
    ):
        if entry["keyword"] not in reasons:
            reasons.append(entry["keyword"])
        if len(reasons) == 3:
            break

    return {
        "category": winner["category"],
        "confidence": determine_confidence(float(winner["score"]), float(runner_up["score"])),
        "reasons": reasons,
    }


def read_baseline_manifest(file_path: Path) -> list[dict[str, Path | str]]:
    if not file_path.exists():
        return []
    entries = json.loads(file_path.read_text(encoding="utf-8"))
    baseline_entries: list[dict[str, Path | str]] = []
    for entry in entries:
        sheet_key = str(entry.get("sheet") or "")
        csv_name = Path(str(entry.get("csv") or "")).name
        tab_name = BASELINE_SHEET_NAMES.get(sheet_key)
        if tab_name and csv_name:
            baseline_entries.append({"tabName": tab_name, "csvPath": EXPORT_DIR / csv_name})
    return baseline_entries


def read_invoice_rows(file_path: Path) -> list[list[str]]:
    records = read_json(file_path, [])
    headers = [
        "source_order_id",
        "source_file",
        "invoice_number",
        "order_number",
        "invoice_date",
        "seller_name",
        "seller_gstin",
        "bill_to_name",
        "ship_to_address",
        "item_description",
        "hsn",
        "quantity",
        "product_rate",
        "discount_percent",
        "taxable_amount",
        "cgst_rate",
        "sgst_rate",
        "cgst_amount",
        "sgst_amount",
        "cess_rate",
        "cess_amount",
        "line_total_amount",
        "invoice_value",
        "parse_quality",
    ]
    rows = [headers]
    for record in records:
        row = [clean_cell(record.get(header)) for header in headers]
        row[0] = clean_cell(record.get("source_order_id") or source_order_id_from_file_name(record.get("source_file")))
        rows.append(row)
    return rows


def read_reconciliation(file_path: Path) -> list[list[str]]:
    payload = read_json(file_path, {"summary": {}, "rows": []})
    summary = payload.get("summary") or {}
    rows = [
        ["Metric", "Value"],
        ["Started From", clean_cell(summary.get("startedFrom"))],
        ["Total Orders Discovered", clean_cell(summary.get("total_orders_discovered"))],
        ["Total Orders In Scope", clean_cell(summary.get("total_in_scope_orders"))],
        ["Total Downloaded", clean_cell(summary.get("total_downloaded"))],
        ["Total Already Downloaded", clean_cell(summary.get("total_already_downloaded"))],
        ["Total Missing Button", clean_cell(summary.get("total_missing_button"))],
        ["Total No Download", clean_cell(summary.get("total_no_download"))],
        ["Total Errors", clean_cell(summary.get("total_error"))],
        ["Total HTML Fallback Captured", clean_cell(summary.get("total_html_fallback_captured"))],
        ["Total Unique Parsed Order IDs", clean_cell(summary.get("total_unique_parsed_order_ids"))],
        ["Dataset Complete", clean_cell(summary.get("dataset_complete"))],
        ["Data Capture Complete", clean_cell(summary.get("data_capture_complete"))],
        [],
        [
            "Order ID",
            "Order Date",
            "Order Amount",
            "Status",
            "Download Status",
            "Invoice Values",
            "Invoice Numbers",
            "Parsed Order Numbers",
            "Amount Match",
            "HTML Capture",
            "HTML Order Number",
            "HTML Bill Summary",
            "Notes",
        ],
    ]
    for row in payload.get("rows") or []:
        rows.append(
            [
                clean_cell(row.get("order_id")),
                clean_cell(row.get("order_date_iso")),
                clean_cell(row.get("order_amount_display")),
                clean_cell(row.get("reconciliation_status")),
                clean_cell(row.get("download_status")),
                clean_cell(row.get("parsed_invoice_values")),
                clean_cell(row.get("parsed_invoice_numbers")),
                clean_cell(row.get("parsed_order_numbers")),
                clean_cell(row.get("amount_match")),
                clean_cell(row.get("html_capture_status")),
                clean_cell(row.get("html_order_number")),
                clean_cell(row.get("html_bill_summary_text")),
                clean_cell(row.get("notes")),
            ]
        )
    return rows


def read_download_summary(file_path: Path) -> list[list[str]]:
    payload = read_json(file_path, {"results": []})
    rows = [
        ["Metric", "Value"],
        ["Started From", clean_cell(payload.get("startedFrom"))],
        ["Orders Discovered", clean_cell(payload.get("ordersDiscovered"))],
        ["Orders Considered", clean_cell(payload.get("ordersConsidered"))],
        ["Orders Targeted This Run", clean_cell(payload.get("ordersPending"))],
        ["Already Downloaded", clean_cell(payload.get("alreadyDownloaded"))],
        ["Downloaded", clean_cell(payload.get("downloaded"))],
        ["Missing Button", clean_cell(payload.get("missingButton"))],
        ["No Download", clean_cell(payload.get("noDownload"))],
        ["Error Count", clean_cell(payload.get("errorCount"))],
        ["HTML Captured", clean_cell(payload.get("htmlCaptured"))],
        [],
        ["Status", "Order ID", "Order Page", "Order Text", "Saved File", "HTML Capture", "HTML JSON"],
    ]
    for result in payload.get("results") or []:
        rows.append(
            [
                clean_cell(result.get("status")),
                clean_cell(result.get("order_id")),
                clean_cell(result.get("href")),
                clean_cell(result.get("text")),
                clean_cell(result.get("file")),
                clean_cell(result.get("html_capture_status")),
                clean_cell(result.get("html_json_path")),
            ]
        )
    return rows


def read_html_fallbacks(file_path: Path) -> list[list[str]]:
    records = read_json(file_path, [])
    headers = [
        "order_id",
        "order_url",
        "order_date_iso",
        "order_amount_display",
        "order_status_text",
        "html_capture_status",
        "html_capture_source",
        "html_order_number",
        "html_items_text",
        "html_bill_summary_text",
        "html_delivery_address_text",
        "html_order_placed_at_text",
        "html_json_path",
        "html_html_path",
    ]
    return [headers, *[[clean_cell(record.get(header)) for header in headers] for record in records]]


def group_invoice_rows_by_order_id(invoice_records: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    rows_by_order_id: dict[str, list[dict[str, Any]]] = {}
    for record in invoice_records or []:
        order_id = record.get("source_order_id") or source_order_id_from_file_name(record.get("source_file"))
        if not order_id:
            continue
        rows_by_order_id.setdefault(order_id, []).append(record)
    return rows_by_order_id


def order_state_label(row: dict[str, Any]) -> str:
    status_text = str(row.get("order_status_text") or "")
    if re.search("cancelled", status_text, re.IGNORECASE) or row.get("reconciliation_status") == "cancelled_order":
        return "Cancelled"
    if re.search("delivered", status_text, re.IGNORECASE):
        return "Delivered"
    return status_text or "Unknown"


def invoice_status_label(row: dict[str, Any]) -> str:
    if row.get("download_file") or row.get("parsed_source_files") or row.get("parsed_invoice_numbers"):
        return "Invoice available"
    if row.get("html_capture_status") or row.get("html_json_path") or row.get("html_html_path"):
        return "HTML fallback"
    return "Missing invoice"


def read_split_ready_rows() -> list[list[str]]:
    reconciliation = read_json(OUTPUTS_DIR / "zepto_reconciliation.json", {"rows": []})
    invoice_records = read_json(OUTPUTS_DIR / "zepto_invoice_rows.json", [])
    annotations = read_json(OUTPUTS_DIR / "zepto_review_annotations.json", {"orders": {}, "lineItems": {}})
    invoice_rows_by_order_id = group_invoice_rows_by_order_id(invoice_records)
    reconciliation_rows_by_order_id = {
        str(row.get("order_id") or ""): row for row in reconciliation.get("rows") or []
    }
    suggested_categories_by_order_id = {
        order_id: suggest_order_category({"invoice_rows": rows}).get("category") or ""
        for order_id, rows in invoice_rows_by_order_id.items()
    }
    headers = [
        "Order Date",
        "Order ID",
        "Line Item Key",
        "Invoice Number",
        "Invoice Available",
        "Order State",
        "Cancelled",
        "Invoice Status",
        "Category",
        "Split Type",
        "Split With",
        "Item Name",
        "Line Total",
        "Ready For Splitwise",
        "Tag Source",
        "Review Status",
        "Notes",
    ]
    rows = [headers]
    line_index_by_order_id: dict[str, int] = {}

    for record in invoice_records:
        order_id = record.get("source_order_id") or source_order_id_from_file_name(record.get("source_file"))
        order_id = str(order_id or "")
        reconciliation_row = reconciliation_rows_by_order_id.get(order_id, {})
        line_index = line_index_by_order_id.get(order_id, 0)
        line_index_by_order_id[order_id] = line_index + 1
        line_item_key = line_item_key_for(order_id, record, line_index)
        order_annotation = (annotations.get("orders") or {}).get(order_id, {})
        line_annotation = (annotations.get("lineItems") or {}).get(line_item_key, {})
        category_fallback = "" if order_annotation.get("suppress_suggested_category") else suggested_categories_by_order_id.get(order_id, "")
        ready_for_splitwise = (
            bool(line_annotation.get("ready_for_splitwise"))
            if has_own(line_annotation, "ready_for_splitwise")
            else bool(order_annotation.get("ready_for_splitwise"))
        )

        rows.append(
            [
                clean_cell(reconciliation_row.get("order_date_iso") or record.get("invoice_date")),
                clean_cell(order_id),
                clean_cell(line_item_key),
                clean_cell(record.get("invoice_number")),
                "yes",
                clean_cell(order_state_label(reconciliation_row)),
                "yes" if re.search("cancelled", order_state_label(reconciliation_row), re.IGNORECASE) else "no",
                clean_cell(invoice_status_label(reconciliation_row)),
                clean_cell(value_from_annotation(line_annotation, order_annotation, "expense_category", category_fallback)),
                clean_cell(value_from_annotation(line_annotation, order_annotation, "split_type", "")),
                clean_cell(value_from_annotation(line_annotation, order_annotation, "split_with", "")),
                clean_cell(record.get("item_description")),
                clean_cell(record.get("line_total_amount")),
                str(ready_for_splitwise),
                split_tag_source(line_annotation, order_annotation),
                clean_cell(review_status_from_annotation(line_annotation, order_annotation)),
                clean_cell(value_from_annotation(line_annotation, order_annotation, "notes", reconciliation_row.get("notes") or "")),
            ]
        )

    return rows


def write_sheet(workbook: Workbook, title: str, rows: list[list[Any]]) -> Worksheet:
    sheet = workbook.create_sheet(title)
    if not rows:
        return sheet

    for row in rows:
        sheet.append([clean_cell(value) for value in row])

    header_fill = PatternFill(fill_type="solid", fgColor="0F766E")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    sheet.freeze_panes = "A2"

    for column_cells in sheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = max((len(str(cell.value or "")) for cell in column_cells), default=0)
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 42)

    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    return sheet


def build_workbook() -> Path:
    OUTPUTS_DIR.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)

    for baseline in read_baseline_manifest(BASELINE_MANIFEST_PATH):
        csv_path = Path(baseline["csvPath"])
        if csv_path.exists():
            write_sheet(workbook, str(baseline["tabName"]), read_csv(csv_path))

    write_sheet(workbook, "Invoice Rows", read_invoice_rows(OUTPUTS_DIR / "zepto_invoice_rows.json"))
    write_sheet(workbook, "Download Summary", read_download_summary(OUTPUTS_DIR / "zepto_download_summary.json"))
    write_sheet(workbook, "Reconciliation", read_reconciliation(OUTPUTS_DIR / "zepto_reconciliation.json"))
    write_sheet(workbook, "HTML Fallback", read_html_fallbacks(OUTPUTS_DIR / "zepto_html_fallbacks.json"))
    write_sheet(workbook, "Split Ready", read_split_ready_rows())

    workbook.save(WORKBOOK_OUT)
    return WORKBOOK_OUT


def main() -> None:
    workbook_path = build_workbook()
    print(json.dumps({"workbook": str(workbook_path)}))


if __name__ == "__main__":
    main()
